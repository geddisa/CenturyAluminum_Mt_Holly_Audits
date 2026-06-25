import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

FILE_NAME = "Audit Schedule - Internal - LPA.xlsx"
LOG_SHEET = "Audit_Log"

# -----------------------------
# MASTER VALUES (FROM YOUR FILE)
# -----------------------------
AUDITORS = [
    "Freddie Gamble", "Anthony Wall", "Tim Kass", "Bryan Profit",
    "Reggie Coleman", "Miguel Frias", "Jacoby Alston",
    "Art DiFilippo", "Jarvis Adelabu", "Brett Meyer"
]

AREAS = ["Maintenance", "Carbon", "Cast House", "Potline", "Environmental"]

LOCATIONS = [
    "PL 136 / Cruce cleaner","PL Room 1 West","PL Room 2 West","PL Room 3 West","PL Room 4 West",
    "PL Room 1 East","PL Room 2 East","PL Room 3 East","PL Room 4 East",
    "PL Services/ Ore Unloading","PL 138 Repair/Rebuild",
    "CB Coke Tank","CB Pitch Tanks","CB Green Mill","CB Rod Shop","CB Bake Oven North","CB Bake Oven South",
    "CH Moldshop","CH Laboratory","CH Shipping/Scales",
    "CH Casting Pits/Billet Inspection Line","CH Dross Room/Furnaces/Sow Line",
    "CH Homogen Furnaces/Pre-Heat/Cooling",
    "MT Crane Shop","MT Fabrication Shop","MT Rebuild Shop","MT Mobile Shop",
    "MT Utilities Shop","MT Compressor Room","MT 044 Warehouse",
    "ENV 161 East Baghouse","ENV 161 West Baghouse",
    "ENV 162 East Baghouse","ENV 162 West Baghouse",
    "ENV Waste Fluid Area","ENV 138 CAA Building","ENV 261G Bake Oven Scrubber"
]

AUDIT_TYPES = ["HK", "Safe Obs", "PPE", "LOTO", "Mobile Equipment"]

# -----------------------------
# ENSURE LOG SHEET EXISTS
# -----------------------------
def ensure_log_sheet():
    wb = load_workbook(FILE_NAME)

    if LOG_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(LOG_SHEET)

        headers = [
            "Audit ID", "Date Entered", "Week Of", "Auditor Name",
            "Area", "Location", "Audit Type", "Score (%)", "Completed"
        ]

        ws.append(headers)
        wb.save(FILE_NAME)
        print("✅ Audit_Log sheet created")

# -----------------------------
# ADD NEW AUDIT
# -----------------------------
def add_audit():
    print("\n--- Add New Audit ---")

    auditor = input("Auditor Name: ")
    area = input("Area: ")
    location = input("Location: ")
    audit_type = input("Audit Type: ")
    week_of = input("Week of (ex: 6/1/26 - 6/5/26): ")
    score = int(input("Score (1-100): "))
    completed = input("Completed (C/c): ")

    # VALIDATION
    if auditor not in AUDITORS:
        print("❌ Invalid auditor")
        return

    if area not in AREAS:
        print("❌ Invalid area")
        return

    if location not in LOCATIONS:
        print("❌ Invalid location")
        return

    if audit_type not in AUDIT_TYPES:
        print("❌ Invalid audit type")
        return

    if not (1 <= score <= 100):
        print("❌ Score must be 1–100")
        return

    if completed.lower() != "c":
        print("❌ Completed must be 'C' or 'c'")
        return

